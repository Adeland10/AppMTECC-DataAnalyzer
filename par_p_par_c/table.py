#coding: utf-8

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Border, Side, PatternFill, Font
from openpyxl.utils import get_column_letter


def table_delta_base(aggregated_conditions):
    rows_base = []
    index = []
    delta_names = ['ΔAmi', 'ΔFsk/IBMX', 'ΔVX770', 'ΔApi', 'ΔInh', 'ΔATP']
    measures = ['GT', 'PD', 'Ieq', 'Iraw', 'RT']

    for patient, conditions in aggregated_conditions.items():
        for (cond1, cond2), deltas in conditions.items():
            row_base = []
            for measure in measures:
                for delta in delta_names:
                    if delta in deltas and measure in deltas[delta]:
                        row_base.append(deltas[delta][measure])
                    else:
                        row_base.append(np.nan)  # Add NaN if the delta does not exist
            rows_base.append(row_base)
            index.append((patient, cond1, cond2))

    columns_base = pd.MultiIndex.from_product([measures, delta_names], names=['Measure', 'Delta'])
    delta_table_base = pd.DataFrame(rows_base, index=pd.MultiIndex.from_tuples(index, names=['Patient', 'Condition1', 'Condition2']), columns=columns_base)

    return delta_table_base


def table_delta_calculated(aggregated_conditions):
    rows_calculated = []
    index = []
    calculated_delta_names = ['ΔFsk + ΔVX770', 'ΔFsk + ΔVX770 + ΔApi', 'ΔAmi', 'ΔInh']
    measures = ['GT', 'PD', 'Ieq', 'Iraw', 'RT']
    measures_rates = ['GT', 'Ieq', 'Iraw']

    for patient, conditions in aggregated_conditions.items():
        for (cond1, cond2), deltas in conditions.items():
            row_calculated = []
            for measure in measures:
                delta_ami = deltas.get('ΔAmi', {}).get(measure, np.nan)
                delta_fsk_ibmx = deltas.get('ΔFsk/IBMX', {}).get(measure, np.nan)
                delta_inh = deltas.get('ΔInh', {}).get(measure, np.nan)
                delta_vx770 = deltas.get('ΔVX770', {}).get(measure, np.nan)
                delta_api = deltas.get('ΔApi', {}).get(measure, np.nan)

                delta_fsk_vx770 = np.nan if np.isnan(delta_fsk_ibmx) or np.isnan(delta_vx770) else delta_fsk_ibmx + delta_vx770
                delta_fsk_vx770_api = np.nan if np.isnan(delta_fsk_vx770) or np.isnan(delta_api) else delta_fsk_vx770 + delta_api

                row_calculated.extend([delta_ami, delta_fsk_ibmx, delta_fsk_vx770, delta_fsk_vx770_api, delta_inh])

            rows_calculated.append(row_calculated)
            index.append((patient, cond1, cond2))

    columns_calculated = pd.MultiIndex.from_product([measures, ['ΔAmi', 'ΔFsk/IBMX', 'ΔFsk + ΔVX770', 'ΔFsk + ΔVX770 + ΔApi', 'ΔInh']], names=['Measure', 'Delta'])
    delta_table_calculated = pd.DataFrame(rows_calculated, index=pd.MultiIndex.from_tuples(index, names=['Patient', 'Condition1', 'Condition2']), columns=columns_calculated)
    
    # Ajouter les colonnes pour les taux
    for measure in measures_rates:
        for delta in calculated_delta_names:
            delta_table_calculated[(measure, f'{delta} T/C')] = np.nan
            delta_table_calculated[(measure, f'{delta} (T-C)/N')] = np.nan
            delta_table_calculated[(measure, f'{delta} T/N')] = np.nan
            delta_table_calculated[(measure, f'{delta} C/N')] = np.nan

    # Réorganiser les colonnes
    delta_table_calculated = reorder_columns(delta_table_calculated)

    return delta_table_calculated


# Fonction pour réorganiser les colonnes par mesure
def reorder_columns(df):
    measures = ['GT', 'Ieq', 'Iraw']
    calculated_deltas = ['ΔFsk + ΔVX770', 'ΔFsk + ΔVX770 + ΔApi']
    rate_types = ['T/C', '(T-C)/N', 'T/N', 'C/N']

    new_columns = []
    for measure in measures:
        # Ajouter les colonnes de base
        # Ajouter ΔAmi et ses taux
        new_columns.append((measure, 'ΔAmi'))
        for rate in rate_types:
            new_columns.append((measure, f'ΔAmi {rate}'))

        #Ajouter ΔFsk/IBMX
        new_columns.append((measure, 'ΔFsk/IBMX'))

        # Ajouter les colonnes calculées et les taux correspondants
        for delta in calculated_deltas:
            new_columns.append((measure, delta))
            for rate in rate_types:
                new_columns.append((measure, f'{delta} {rate}'))
            
        # Ajouter ΔInh et ses taux
        new_columns.append((measure, 'ΔInh'))
        for rate in rate_types:
            new_columns.append((measure, f'ΔInh {rate}'))

    return df[new_columns]

# Fonction pour ajuster la largeur des colonnes
def adjust_column_width(filepath):
    wb = load_workbook(filepath)

    colors = {
        'GT': 'AFEEEE',    # turquoise clair
        'PD': 'FFDAB9',  # orange clair
        'Ieq': 'E6E6FA',   # Violet clair
        'Iraw': '98FB98',  # menthe clair
        'RT': 'F3E5AB',  # vanille
    }

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        for col in sheet.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) *1.2
            column = col[0].column_letter if not isinstance(col[0], MergedCell) else col[0].coordinate[:1]  # Get the column name
            sheet.column_dimensions[column].width = adjusted_width
            
            # Ajuster la hauteur de la première ligne pour les titres des mesures
        sheet.row_dimensions[1].height = 30  # Augmenter la hauteur à 30 (en points)
        
            # Ajuster l'épaisseur de la ligne des titres
        for row in sheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.border = Border(bottom=Side(style='medium'))
                measure = cell.value.split()[0] if cell.value else None
                if measure in colors:
                    fill = PatternFill(start_color=colors[measure], end_color=colors[measure], fill_type="solid")
                    cell.fill = fill

    wb.save(filepath)

